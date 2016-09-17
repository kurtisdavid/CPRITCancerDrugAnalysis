# Goal of this file: to collect all of the "Preferred Terms" in the MedDRA ontology
# In an earlier version (MedDRA_ADR_old.py) I did not paginate through the BioPortal REST API, so I had missed out on some of the terms
# This ensures I have all of the ones I need for the desired System Organ classes.


from urllib2 import Request, urlopen, URLError
import urllib
from openpyxl import load_workbook, Workbook
import json

def main():

	# loads database of adverse drug events in the saved Excel file
	# can replace these with any type of database needed
	wb = load_workbook("ADE_Database.xlsx")
	ws = wb["Sheet1"]

	MedDRA = {}

	for i in range(2, 21071):

		if ws["A" + str(i)].value not in MedDRA:

			MedDRA[ws["A" + str(i)].value] = []

		if ws["B" + str(i)].value not in MedDRA[ws["A" + str(i)].value]:

			MedDRA[ws["A" + str(i)].value].append(ws["B" + str(i)].value)

	index = [21071]

	# used to search each ADE in the MedDRA ontology
	# sends information to the recurse() method that 
	for SOC in MedDRA:

		print(SOC)

		orig = SOC

		ids = []

		ADRs = MedDRA[SOC]

		orig = orig.replace(" ", "%20")

		url = "http://data.bioontology.org/search?q=" + orig +"&ontologies=MEDDRA&apikey=3cb45b79-b7dd-42ab-8f44-394434267e1b"
		request = Request(url)
		response = urlopen(request)
		report = response.read()
		parsed_json = json.loads(report)
		ids.append(parsed_json["collection"][0]["@id"].split("/")[-1])

		

		new_url = parsed_json["collection"][0]["links"]["children"]

		recurse(new_url,ids, ADRs, index, ws, SOC)
		print("Saving...")
		wb.save("ADE_Database_ok.xlsx")
		wb = load_workbook("ADE_Database_ok.xlsx")
		ws = wb["Sheet1"]
	

# very similar algorithm to old version, however, adds a paginating capability to make sure every leaf is found.
# can check the comments of MedDRA_ADR_old.py for specifics on the tree traversion
def recurse(link, m_ids, ADRs, index, ws, SOC):

	ids = m_ids[:]


	url = link
	request = Request(url, headers= {"Authorization": "apikey token=3cb45b79-b7dd-42ab-8f44-394434267e1b"})
	response = urlopen(request)
	report = response.read()
	parsed_json = json.loads(report)

	collection = parsed_json["collection"]

	while parsed_json["pageCount"]>1 and parsed_json["links"]["nextPage"]:

		url = parsed_json["links"]["nextPage"]
		print(url)
		request = Request(url, headers= {"Authorization": "apikey token=3cb45b79-b7dd-42ab-8f44-394434267e1b"})
		response = urlopen(request)
		report = response.read()
		parsed_json = json.loads(report)

		for i in range(len(parsed_json["collection"])):

			collection.append(parsed_json["collection"][i])

	if collection==[]:

		correct_url = "http://data.bioontology.org/ontologies/MEDDRA/classes/http%3A%2F%2Fpurl.bioontology.org%2Fontology%2FMEDDRA%2F" + str(ids[-2]) + "/children"
		correct_request = Request(correct_url, headers= {"Authorization": "apikey token=3cb45b79-b7dd-42ab-8f44-394434267e1b"})
		correct_response = urlopen(correct_request)
		correct_report = correct_response.read()
		correct_parsed = json.loads(correct_report)

		collection = correct_parsed["collection"]

		while correct_parsed["links"]["nextPage"]:

			correct_url = correct_parsed["links"]["nextPage"]
			correct_request = Request(correct_url, headers= {"Authorization": "apikey token=3cb45b79-b7dd-42ab-8f44-394434267e1b"})
			correct_response = urlopen(correct_request)
			correct_report = correct_response.read()
			correct_parsed = json.loads(correct_report)

			for i in range(len(correct_parsed["collection"])):

				collection.append(correct_parsed["collection"][i])

		for i in range(len(collection)):

			if collection[i]["prefLabel"] not in ADRs:

				ADR = collection[i]["prefLabel"]

				print(SOC + ":" + ADR)

				ADRs.append(ADR)

				ws["A" + str(index[0])].value = SOC
				ws["B" + str(index[0])].value = ADR

				index[0] += 1

			else:

				print(collection[i]["prefLabel"])

		print(len(ADRs))


	else:

		names = []

		for i in range(len(collection)):

			names.append(collection[i]["prefLabel"])

		for i in range(len(collection)):

			if not set(names)<=set(ADRs):

				new_ids = m_ids[:]
				new_ids.append(collection[i]["@id"].split("/")[-1])

				new_url = parsed_json["collection"][i]["links"]["children"]

				recurse(new_url, new_ids, ADRs, index, ws, SOC)

			else:

				break

	return ADRs
			






main()