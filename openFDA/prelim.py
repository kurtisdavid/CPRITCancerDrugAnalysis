# goal: to prep the original JSON into only what we need
# main thing to check is the # reports for the relationship must be > 10 for the study
# eventually creates 2 new JSON files, one where the keys are ADEs (reactions) and the other where the keys are drugs 

# many versions of this existed, due to changes of accumulate.py; this should reflect the final conditions we came up with

import simplejson as json
from collections import OrderedDict
from openpyxl import load_workbook, Workbook
from urllib2 import Request, urlopen, URLError

# filters through the reaction-drug relationships to only those with counts greater than 10
# also only uses those that have ACME drugs within them (molecules within our cancer cell line dataset)
def main():

	file = open("myt_acme_FINAL.json")
	json_data = json.load(file)
	file.close()

	wb = load_workbook("ADE_Database.xlsx")
	ws = wb["Sheet1"]

	syn_dict = {}

        # initializes the ADE database collected initially (from MedDRA)
	for i in range(2, 22046):

		ADE = str(ws["B" + str(i)].value).upper().replace("'", "^")
		if ADE not in syn_dict:

			syn_dict[ADE] = ADE


		classifies = ws["C" + str(i)].value

		if classifies is None:

			continue

		syns = classifies.split("$")

		for syn in syns:

			use = syn.upper().replace("'", "^")

			if use not in syn_dict:

				syn_dict[use] = ADE


	real_dict = {}
	summ = 0

	for ADE in json_data:

		if ADE not in syn_dict:

			continue

		real = syn_dict[ADE]

		if real not in real_dict:

			real_dict[real] = {}

		drugs = json_data.get(ADE, {})

		for drug in drugs:

			if drug not in real_dict[real]:

				real_dict[real][drug] = drugs[drug]

			else:

				real_dict[real][drug] += drugs[drug]


	final_dict = {}

	for ADE in real_dict:

		final_dict[ADE] = {}

		drugs = real_dict[ADE]

		for drug in drugs:

			if drugs[drug] >= 10:

				final_dict[ADE][drug] = drugs[drug]

				summ += drugs[drug]

	summ += json_data.get(ADE,"")[drug]

	f = open("myt_acme_FINAL_filtered.json", "w")
	json.dump(final_dict, f, indent = 4, sort_keys=True)
	print(summ)

# reorders the JSON files in alphabetical order of the ADEs
def changeOrder():

	file = open("myt_acme_FINAL_filtered.json")
	json_data = json.load(file)
	file.close()

	file = open("unused3.txt")

	notUsed = []

	for line in file:

		notUsed.append(line[:-1])

	orig_dict = {}

	for ADE in json_data:

		if ADE.replace("^","'") in notUsed:

			continue

		attempt = {}

		for drug in json_data[ADE]:

			if json_data[ADE][drug] >= 10:

				attempt[drug] = json_data[ADE][drug]

		if attempt != {}:

			orig_dict[ADE] = attempt

	new = []

	lol_dict = {}

	for ADE in orig_dict:

		current = orig_dict[ADE]
		lol_dict[ADE] = OrderedDict([(k,v) for v,k in sorted([(v,k) for k,v in current.items()], reverse=True)])

	lol = lol_dict.items()

	lol.sort(key=lambda x:x[0])

	final = OrderedDict(lol)

	f = open("ADE_Drug_ACME_Update_FINAL.json", "w")
	json.dump(final, f, indent=4)

# counts the total number of reports within a SOC
# each SOC contains many ADEs, and this is needed for our Proportional Reporting Ratios
def aggregateADE():

	file = open("ADE_Drug_ACME_Update_FINAL.json")
	json_data = json.load(file)
	file.close()

	old_dict = {}

	for ADE in json_data:

		correct_ADE = ADE.replace("^","'")

		old_dict[correct_ADE] = {}

		for drug in json_data[ADE]:

			correct_drug = drug.replace("^", "'")

			old_dict[correct_ADE][correct_drug] = json_data[ADE][drug]

	new_dict = {}

	ADE_db = load_workbook("ADE_Database.xlsx")
	ws = ADE_db["Sheet1"]

	SOC_ADE = {}

	for i in range(2,22046):

		PT = ws["B" + str(i)].value.upper()
		SOC = ws["A" + str(i)].value
		
		if SOC not in SOC_ADE:

			SOC_ADE[SOC] = [PT]

		else:

			SOC_ADE[SOC].append(PT)

		others = ""

		if ws["C" + str(i)].value:

			others = ws["C" + str(i)].value

		else:

			continue

		terms = others.split("$")

		for term in terms:

			if term == "":

				continue

			SOC_ADE[SOC].append(term.upper())


	SOC_counts = {}

	record = open("SOC_counts_FINAL2.txt", "w")

	check = open("unused_new_FINAL.txt", "w")

	# for SOC in SOC_ADE:

	# 	SOC_counts[SOC] = 0

	# 	for ADE in SOC_ADE[SOC]:

	# 		if ADE.upper() not in old_dict:

	# 			check.write(SOC + "$" + str(ADE.upper()) + "\n")
	# 			continue

	# 		current = old_dict[ADE.upper()]

	# 		for drug in current:

	# 			SOC_counts[SOC] += current[drug]

	# 	print(SOC + " " + str(SOC_counts[SOC]))
	# 	record.write(SOC + " " + str(SOC_counts[SOC]) + "\n")

	for SOC in SOC_ADE:

		SOC_counts[SOC] = 0

	for ADE in old_dict:

		found = False

		for SOC in SOC_ADE:

			if ADE in SOC_ADE[SOC]:

				found = True

				for drug in old_dict[ADE]:

					SOC_counts[SOC] += old_dict[ADE][drug]

		if not found:

			check.write(ADE + "\n")

	for SOC in SOC_counts:

		print(SOC + " " + str(SOC_counts[SOC]))
		record.write(SOC + " " + str(SOC_counts[SOC]) + "\n") 



	record.close()

# creates another JSON that is ordered by drug names (alphabetically)
def toDrug():

	file = open("ADE_Drug_ACME_Update_FINAL.json")
	json_data = json.load(file)
	file.close()

	new_dict = {}

	for ADE in json_data:

		for drug in json_data[ADE]:

			if drug not in new_dict:

				new_dict[drug] = {}

			current = new_dict[drug]

			if ADE not in current:

				current[ADE] = json_data[ADE][drug]

			else:

				current[ADE] += json_data[ADE][drug]

	lol_dict = {}

	for drug in new_dict:

		current = new_dict[drug]
		lol_dict[drug] = OrderedDict([(k,v) for v,k in sorted([(v,k) for k,v in current.items()], reverse=True)])

	lol = lol_dict.items()

	lol.sort(key=lambda x:x[0])

	final = OrderedDict(lol)

	f = open("Drug_ADE_ACME_Update_FINAL.json", "w")
	json.dump(final, f, indent=4)

# checked the terms not within the database; realized that I hadn't paginated through BioPortal (fixed in another loop through of MedDRA_ADR.py)
def check():

	errors = open("unused3.txt")

	real_errors = open("checked.txt","w")

	wb = load_workbook("ADE_Database.xlsx")
	ws = wb["Sheet1"]

	ADE_list = {}

	for i in range(2, 22046):

		current = []

		current.append(ws["B" + str(i)].value)

		if not ws["C" + str(i)].value:

			continue

		nonPT = ws["C" + str(i)].value.split("$")

		for term in nonPT:

			current.append(term)

		for ADE in current:

			if ADE not in ADE_list:

				ADE_list[ADE] = ""

	count = 0
	for line in errors:

		ADE = line[:-1]

		if ADE not in ADE_list:

			real_errors.write(line)
			print(ADE)
			count+=1

	print(count)

# just combined the 2 SOCS into 1 to be able to compare results to another study
def combine():

	wb = load_workbook("ADE_Database.xlsx")
	ws = wb["Sheet1"]

	for i in range(2,22046):

		SOC = ws["A" + str(i)].value

		if SOC == "Infections and infestations" or SOC == "Immune system disorders":

			ws["A" + str(i)].value = "Infections, infestations and immune system disorders"

	wb.save("ADE_Database_new.xlsx")

# QOL to make the SOC database in alpha order
def sortExcel():

	wb = load_workbook("ADE_Database_new.xlsx")
	ws = wb["Sheet1"]

	SOC_row = {}

	for i in range(2,22046):

		SOC = ws["A" + str(i)].value
		ADE = ws["B" + str(i)].value
		classifies = ws["C" + str(i)].value

		if not classifies:

			classifies = ""

		if SOC not in SOC_row:

			SOC_row[SOC] = [[ADE, classifies]]

		else:

			SOC_row[SOC].append([ADE, classifies])


	sort = OrderedDict(sorted([(k,v) for (k,v) in SOC_row.items()]))

	index = 2

	for SOC in sort.keys():

		rows = sort[SOC]

		for row in rows:

			ws["A" + str(index)] = SOC
			ws["B" + str(index)] = row[0]
			ws["C" + str(index)] = row[1]

			index += 1

	wb.save("ADE_Database_sorted.xlsx")

# main()
# changeOrder()
# aggregateADE()
# toDrug()
# check()
# combine()
# sortExcel()
