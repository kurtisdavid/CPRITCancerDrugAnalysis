#!/usr/bin/env python

''' Simple example of reading all of the zip files from the openFDA download
    and doing something with them. In this case, we are building an index
    of all the medicinalproduct and drugcharacterization values and a count
    of how often each occurred.

    template from: https://gist.github.com/HansNelsen/aeec93279dcd1792855d39fc37bead2e
'''

# used its initial structure and modified the conditions to match what the research needed

from collections import defaultdict
import glob
import simplejson as json
from os.path import basename
import re
from openpyxl import load_workbook, Workbook

acme_syns = {}
wb = load_workbook("acmeSynonyms.xlsx")
ws = wb["Sheet1"]

ADE_syns = {}
wb_M = load_workbook("ADE_Database.xlsx")
ws_M = wb_M["Sheet1"]

# initialize ACME synonyms
for i in range(2,22997):

	acme = str(ws["A" + str(i)].value)

	if acme not in acme_syns:

		acme_syns[acme] = [acme.upper()]

	syn = str(ws["C" + str(i)].value).upper()

	if syn not in acme_syns[acme]:

		acme_syns[acme].append(syn)

original = [acme for acme in acme_syns]

# kept track of every synonym of the drugs we were looking at
for acme in original:

	syns = acme_syns[acme]

	for syn in syns:

		if syn not in acme_syns:

			acme_syns[syn] = acme_syns[acme]

# also created a list for each ADE synonym that contained all of its other synonyms, so that it would be easier to search if a certain term had been used
# this is on the condition that a synonym is the same as the original Preferred Term
for i in range(2,22046):

	PT = ws_M["B" + str(i)].value

	if PT in ADE_syns:

		continue

	ADE_syns[PT.upper()] = [PT.upper()]
	current = ADE_syns[PT.upper()]

	lower = ""

	if ws_M["C" + str(i)].value is None:

		continue

	lower = ws_M["C" + str(i)].value

	syns = lower.split("$")

	for syn in syns:

		if syn.upper() not in current:

			current.append(syn.upper())

	for syn in syns:

		if syn.upper() not in ADE_syns:

			ADE_syns[syn.upper()] = current


DATA_FILES = glob.glob('./*/*.json')

reaction_drug_counts = {}

# goes through every JSON file downloaded from openFDA
for filename in DATA_FILES:

	print(filename)
  
	file = open(filename)
	json_data = json.load(file)
	file.close()

	for row in json_data['results']:

		done = []

                # goes through the reactions observed in patient
                #       each loop goes through the drugs given, so that each relationship is recorded per report
                # makes sure that if the same reaction (as in a synonym or repeats) has been looped already, goes to the next 
		for reaction in row.get('patient', {}).get('reaction', []):

			MedDRA_term = reaction.get('reactionmeddrapt','unknown').upper()

			if MedDRA_term in done:

				continue

			if MedDRA_term not in reaction_drug_counts:

				reaction_drug_counts[MedDRA_term] = {}

			done.append(MedDRA_term)

			found = []

                        # goes through all the drugs taken (main drug ingredients, substance list)
                        # gets rid of misread data (such as only characters)
			for drug in row.get('patient', {}).get('drug', []):

				characterization = int(drug.get('drugcharacterization', '0'))

				# makes sure it is a suspect drug
				if characterization == 1:

					# check indication!!!

					drug_indication = drug.get("drugindication", "000")

					# makes sure that if the drug indication is a synonym of the current reaction, don't count this drug (since it's not an actual side effect)

					if drug_indication != "000":

						reason = drug_indication.upper()

						if (MedDRA_term in ADE_syns and reason in ADE_syns[MedDRA_term]):

							continue

					# checks if active substance is an acme molecule

					active_substance = drug.get("activesubstance", "000")

					if active_substance != "000":

						name = active_substance.get("activesubstancename", "lol")

						if name != "lol":

							nameU = name.upper()

							if nameU in found:

								continue

							if nameU in acme_syns:

								found.append(nameU)
								check = acme_syns[nameU]

								for syn in check:

									found.append(syn)

								current_counts = reaction_drug_counts[MedDRA_term]

								if nameU not in current_counts:

									current_counts[nameU] = 1

								else:

									current_counts[nameU] += 1

					# check to see if acme molecules are in the main ingredient list

					openfda = drug.get('openfda', {})

					if openfda != {}:

						substances = openfda.get("substance_name", [])

						for substance in substances:

							component = substance.upper()

							if component in found:

								continue

							if component in acme_syns:

								found.append(component)

								check = acme_syns[component]

								for syn in check:

									found.append(syn)

								current_counts = reaction_drug_counts[MedDRA_term]

								if component not in current_counts:

									current_counts[component] = 1

								else:

									current_counts[component] += 1

					product = drug.get('medicinalproduct', 'unknown')

					if not product or re.search("^[\W_]+$", product):

						continue

					product = product.upper()

					if product in found:

						continue

					if product in acme_syns:

						check = acme_syns[product]

						for syn in check:

							found.append(syn)

					current_counts = reaction_drug_counts[MedDRA_term]

					if product not in current_counts:

						current_counts[product] = 1

					else:

						current_counts[product] += 1

					found.append(product)



# after collecting the # reports each reaction-drug relationship has been found, dumps all into a new JSON
# effectively reduces the information to loop through from 200GB to 200MB
f = open("myt_acme_FINAL.json","w")
json.dump(reaction_drug_counts,f, indent = 4, sort_keys=True)
