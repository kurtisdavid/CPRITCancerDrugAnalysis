# Goal: to see how well the models could predict signals of the other drugs not found in the FDA yet
# results: many are found to have a negative signal, however, this is due to such a small dataset (of only 80 points) in comparison to the # of total features (~642 cell lines)

from openpyxl import load_workbook, Workbook
import numpy as np
from scipy import interp
from scipy.sparse import csr_matrix
from sklearn import linear_model, cross_validation, feature_selection
from sklearn.metrics import roc_curve, auc

def main():

	# formatTest()
	predict()


def formatTest():

	dimension_wb = load_workbook("dimensionMatrix.xlsx")
	test_ws = dimension_wb["test"]
	train_ws1 = dimension_wb["S1"]

	data_wb = load_workbook("ACMEsetWithCIDs_new.xlsx")
	cpd_ws = data_wb["S1"]
	AUC_ws = data_wb["S3"]

	used_cpd = []

	ccl_values = {}

	ccl_normals = {}
	test_cpds = {}

	print("0")

	for i in range(2, 44758):

		cpd = train_ws1["A" + str(i)].value
		AUC = train_ws1["D" + str(i)].value

		if cpd not in used_cpd:

			used_cpd.append(cpd)

		ccl = train_ws1["C" + str(i)].value

		if ccl not in ccl_values:

			ccl_values[ccl] = {cpd:AUC}

		else:

			ccl_values[ccl][cpd] = AUC

	print("1")

	for ccl in ccl_values:

		current = ccl_values[ccl]

		AUCs = [v for k,v in current.items()]

		mu = np.mean(AUCs)
		sigma = np.std(AUCs)

		ccl_normals[ccl] = [mu, sigma]

	print("2")

	for i in range(2, 483):

		cpd = cpd_ws["B" + str(i)].value

		if cpd not in used_cpd:

			test_cpds[cpd] = {}

	print("3")

	for i in range(2, 260498):

		cpd = AUC_ws["A" + str(i)].value
		ccl = AUC_ws["B" + str(i)].value
		AUC = float(AUC_ws["E" + str(i)].value)

		if cpd not in test_cpds:

			continue

		if ccl not in ccl_normals:

			continue

		mu = ccl_normals[ccl][0]
		sigma = ccl_normals[ccl][0]

		test_cpds[cpd][ccl] = (AUC-mu)/sigma

	print("4")

	for cpd in test_cpds:

		for ccl in ccl_normals:

			if ccl not in test_cpds[cpd]:

				test_cpds[cpd][ccl] = 0

	ind = 2

	print("5")

	for cpd in test_cpds:

		current = test_cpds[cpd]

		for ccl in current:

			test_ws["A" + str(ind)].value = cpd
			test_ws["B" + str(ind)].value = ccl
			test_ws["C" + str(ind)].value = current[ccl]
			ind += 1

	dimension_wb.save("dimensionMatrix.xlsx")

def predict():

	wb = load_workbook("dimensionMatrix.xlsx")
	test_ws = wb["test"]
	ws1 = wb["S1"]
	ws2 = wb["S2"]
	predict_ws = wb["predict"]


	letters = {"C": 558, "D": 348, "E": 47, "F": 103, "G": 63, "H": 14, "I": 13, "J": 7, "K": 167, "L": 58, "M": 64, "N": 58, "O": 273, "P": 10, "Q": 27, "R": 213, "S": 13, "T": 30, "U": 45, "V": 183, "W": 239}

	unique_ccl = []
	for i in range(2,44758):

		ccl = int(ws1["C" + str(i)].value)

		if ccl not in unique_ccl:

			unique_ccl.append(ccl)

	unique_ccl = sorted(unique_ccl)

	cpd_signal = {}

	for i in range(2,82):

		cpd = int(ws2["A" + str(i)].value)

		if cpd not in cpd_signal:

			cpd_signal[cpd] = {}

		for letter in letters:

			cpd_signal[cpd][letter] = ws2[letter + str(i)].value

	cpd_zscores = {}

	for i in range(2,44758):

		cpd = int(ws1["A" + str(i)].value)

		if cpd not in cpd_zscores:

			cpd_zscores[cpd] = csr_matrix((1, 642)).toarray()[0]

		ccl = int(ws1["C" + str(i)].value)
		current = cpd_zscores[cpd]
		z_score = ws1["E" + str(i)].value

		current[unique_ccl.index(ccl)] = z_score

	test_zscores = {}

	for i in range(2,257443):

		cpd = test_ws["A" + str(i)].value

		if cpd not in cpd_zscores and cpd not in test_zscores:

			test_zscores[cpd] = csr_matrix((1,642)).toarray()[0]

		ccl = test_ws["B" + str(i)].value
		current = test_zscores[cpd]
		z_score = test_ws["C" + str(i)].value

		current[unique_ccl.index(ccl)] = z_score

	X = []
	Y = []

	X_test = []


	

	for letter in letters:

		print(letter + " !!! ")
		means = np.zeros(642)

		for cpd in cpd_signal:

			Y.append(cpd_signal[cpd][letter])
			X.append(cpd_zscores[cpd])

		X = np.asarray(X)
		Y = np.asarray(Y)

		clf = linear_model.LogisticRegression()

		cv = cross_validation.LeaveOneOut(80)

		probas_ = []

		for i, (train, test) in enumerate(cv):

			current_x = X[train][:]
			current_y = Y[train][:]

			sel = feature_selection.GenericUnivariateSelect(feature_selection.f_regression, mode='k_best', param = letters[letter])
			sel.fit(current_x, Y[train])
			current_x = sel.transform(current_x)
			means = np.add(means, sel.scores_)

			proba_ = linear_model.LogisticRegression().fit(current_x, current_y).predict_proba(sel.transform(X[test]))[0][1]
			probas_.append(proba_)

		fpr, tpr, threshhold = roc_curve(Y, probas_)
		roc_auc = auc(fpr,tpr)

		means = np.multiply(means, 1/(len(cv)))

		cell_lines = {}

		for i in range(len(means)):

			cell_lines[unique_ccl[i]] = means[i]

		sorted_lines = [(k,v) for v,k in sorted([(v,k) for k,v in cell_lines.items()])]
		
		wanted = [sorted_lines[i] for i in range(len(sorted_lines)-1, len(sorted_lines)-1-letters[letter], -1)]

		indeces = [unique_ccl.index(wanted[i][0]) for i in range(len(wanted))]

		clf_new = linear_model.LogisticRegression()

		transformed_X = []

		for i in range(len(X)):

			current = X[i]
			transformed_X.append([current[j] for j in range(len(current)) if j in indeces])

		clf_new.fit(transformed_X, Y)

		


		X = []
		Y = []
		cpds = []

		for cpd in test_zscores:

			values = test_zscores[cpd]

			new = [values[i] for i in range(len(values)) if i in indeces]
			X_test.append(new)
			cpds.append(cpd)

		X_test = np.asarray(X_test)

		predictions = clf_new.predict(X_test)

		for i in range(len(predictions)):

			ind = i+2

			predict_ws["A" + str(ind)].value = cpds[i]
			predict_ws[letter + str(ind)].value = predictions[i]

			

		X_test = []

	wb.save("dimensionMatrix.xlsx")






main()
