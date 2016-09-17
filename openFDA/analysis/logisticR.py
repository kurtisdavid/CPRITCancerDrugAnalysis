# Goal: Apply ML techniques to classify collected data from openFDA and MedDRA

from __future__ import division
import simplejson as json
from collections import OrderedDict
from openpyxl import load_workbook, Workbook
from urllib2 import Request, urlopen, URLError
import numpy as np
from scipy import interp
from scipy.sparse import csr_matrix
from sklearn import datasets, neighbors, linear_model, cross_validation, feature_selection, svm, grid_search
from sklearn.metrics import roc_curve, auc
from sklearn.feature_selection import RFECV
import matplotlib.pyplot as plt
import math
import matplotlib.mlab as mlab
import time as t
def main():

	# cleanUp()
	# normalize()
	# update()
	# check()
	# practice()
	# matchCIDs()
	# combine()
	# addPRR()
	# correctPRR()
	# binarize()
	# createVector() 
	# signalDistribution()
	# practice()
	# calculateError()
	# plotComparison()
	# cellLineIdentification()
	# mutationCollect()
	# RFECV_select()
	# SOC_mutation()
	# mutation_SOC()
	logRegressionFirst()
	# findDifferences()

# removed the entries which had a PRR of 0 and did not pass the >10 rule
def cleanUp():

	wb = load_workbook("contingency_tables3.xlsx")
	ws1 = wb["Sheet1"]
	ws2 = wb["Sheet2"]

	letters = ["A", "B", "C", "D", "E", "F", "G"]

	index = 2
	for i in range(2, 9830):

		# this means that the drug had no reports greater than 10

		if ws1["E" + str(i)].value==0 and ws1["C" + str(i)].value==0:

			continue

		for letter in letters:

			ws2[letter + str(index)].value = ws1[letter + str(i)].value

		index += 1

	wb.save("contingency_tables3.xlsx")

# normalized AUCs of dataset over every cell line
def normalize():

	wb = load_workbook("ACMEsetWithCIDs_new3.xlsx")
	ws9 = wb["S9"]

	ccl_dict = {}

	for i in range(2, 46697):

		cmpd_index = int(ws9["A" + str(i)].value)
		ccl_index = int(ws9["C" + str(i)].value)
		AUC = ws9["D" + str(i)].value

		if ccl_index not in ccl_dict:

			ccl_dict[ccl_index] = {cmpd_index:AUC}

		else:

			ccl_dict[ccl_index][cmpd_index] = AUC
	
	ccl_normalized = {}

	for ccl in ccl_dict:

		ccl_normalized[ccl] = {}

		current = ccl_dict[ccl]

		AUCs = [[k for k,v in current.items()],[v for k,v in current.items()]]

		mu = np.mean(AUCs[1])
		sigma = np.std(AUCs[1])

		data = [mu, sigma]

		for i in range(len(AUCs[0])):

			AUCs[1][i] = (AUCs[1][i]-mu)/sigma
			ccl_normalized[ccl][AUCs[0][i]] = [AUCs[1][i], data]

	# x = []

	# for ccl in ccl_dict:
	# 	x.append(len(ccl_normalized[ccl]))



	# n, bins, patches = plt.hist(x, 50, normed=1, facecolor='green', alpha=0.75)
	
	# y = mlab.normpdf(bins, np.mean(x), np.std(x))

	# print("Mean: " + str(np.mean(x)))
	# print("Standard Deviation: " + str(np.std(x)))
	# l = plt.plot(bins, y, 'r--', linewidth=1)

	# plt.xlabel('Drugs in CCL')
	# plt.ylabel('Probability')
	# plt.title('Histogram of CCL Sizes')
	# plt.show()

	for i in range(2, 46697):

		cpd_index = int(ws9["A" + str(i)].value)
		ccl_index = int(ws9["C" + str(i)].value)

		if len(ccl_normalized[ccl_index])>50:

			ws9["E" + str(i)].value = ccl_normalized[ccl_index][cpd_index][0]
			ws9["F" + str(i)].value = ccl_normalized[ccl_index][cpd_index][1][0]
			ws9["G" + str(i)].value = ccl_normalized[ccl_index][cpd_index][1][1]

		else:

			ws9["E" + str(i)].value = ""

	wb.save("ACMEsetWithCIDs_new3.xlsx")

# made a new worksheet in excel only with the entries needed
def update():

	wb = load_workbook("ACMEsetWithCIDs.xlsx")
	ws3 = wb["S3"]
	ws8 = wb["S8"]

	triples = {}
	for i in range(2,260498):

		ccl = int(ws8["A" + str(i)].value)
		cpd = int(ws8["B" + str(i)].value)
		normal = ws8["C" + str(i)].value

		if cpd not in triples:

			triples[cpd] = {}

		triples[cpd][ccl] = normal

	for i in range(2,260498):

		cpd = ws3["A" + str(i)].value
		ccl = ws3["B" + str(i)].value

		ws3["D" + str(i)].value = triples[cpd][ccl]

	wb.save("ACMEsetWithCIDs_new.xlsx")

# more fixing of the workbook; trying to make saving data easier
def matchCIDs():

	wb = load_workbook("ACMEsetWithCIDs_new.xlsx")
	ws1 = wb["S1"]
	ws3 = wb["S3"]

	cpd_CID = {}

	for i in range(2,483):

		CID = str(ws1["A" + str(i)].value)
		cpd = int(ws1["B" + str(i)].value)

		cpd_CID[cpd] = CID

	for i in range(2, 260498):

		cpd = int(ws3["A" + str(i)].value)
		ws3["E" + str(i)].value = cpd_CID[cpd]

	wb.save("ACMEsetWithCIDs_new.xlsx")

# final change to the desired columns (basically only obtain the drugs we want and their associated cell lines)
def combine():

	wb1 = load_workbook("contingency_tables3.xlsx")
	ws2 = wb1["Sheet2"]

	wb = load_workbook("ACMEsetWithCIDs_new.xlsx")
	ws_id = wb["S1"]
	ws_data = wb["S3"]
	ws_real = wb["S9"]

	acme_data = {}

	SOC_index = {}

	CID_cpd = {}

	for i in range(2, 23):

		SOC = ws2["B" + str(i)].value
		if SOC not in SOC_index:

			SOC_index[SOC] = i-1

	for i in range(2, 1724):

		acme = ws2["A" + str(i)].value

		if acme not in acme_data:

			acme_data[acme] = {}

		SOC = ws2["B" + str(i)].value
		PRR = ws2["G" + str(i)].value

		acme_data[acme][SOC_index[SOC]] = PRR

	for i in range(2, 483):

		CID = str(ws_id["A" + str(i)].value)
		cpd_name = str(ws_id["C" + str(i)].value)
		cpd_index = ws_id["B" + str(i)].value

		CID_cpd[CID] = [cpd_name, cpd_index]

	letters = ["A", "B", "C", "D", "E", "F", "G", "H", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]


	acme_names = [acme.upper() for acme in acme_data.keys()]

	cpd_ccl = {}
	for CID in CID_cpd:

		if CID_cpd[CID][0].upper() in acme_names:

			cpd_ccl[CID_cpd[CID][1]] = {}

	orig_data = {}

	index = 2
	for i in range(2, 260498):

		cpd = ws_data["A" + str(i)].value

		if cpd not in cpd_ccl:

			continue

		ccl = ws_data["B" + str(i)].value
		CID = ws_data["C" + str(i)].value
		AUC = ws_data["E" + str(i)].value

		ws_real["A" + str(index)].value = cpd
		ws_real["C" + str(index)].value = ccl
		ws_real["B" + str(index)].value = CID
		ws_real["D" + str(index)].value = AUC
		index += 1

	wb.save("ACMEsetWithCIDs_new3.xlsx") 

# add in the PRRs to dimensionMatrix (since this is what is used to store the training dataset)
def addPRR():

	wb = load_workbook("dimensionMatrix.xlsx")
	ws = wb["Sheet1"]
	ws_new = wb["Sheet2"]

	wb_data = load_workbook("ACMEsetWithCIDs_new.xlsx")
	ws1 = wb_data["S1"]

	wb_PRR = load_workbook("contingency_tables1.xlsx")
	ws_PRR = wb_PRR["Sheet1"]

	cpd_CID = {}
	for i in range(2,44598):

		cpd_index = ws["A" + str(i)].value

		if cpd_index not in cpd_CID:

			cpd_CID[cpd_index] = ""

	for i in range(2,483):

		CID = ws1["A" + str(i)].value
		cpd_index = ws1["B" + str(i)].value

		if cpd_index in cpd_CID:

			cpd_CID[cpd_index] = CID

	index = 2
	for cpd in cpd_CID:

		ws_new["A" + str(index)].value = cpd
		ws_new["B" + str(index)].value = cpd_CID[cpd]
		index += 1

	wb.save("dimensionMatrix.xlsx")


def correctPRR():

	wb = load_workbook("dimensionMatrix.xlsx")
	ws = wb["Sheet2"]

	CID_name = {}

	for i in range(2,82):

		if ws["B" + str(i)].value not in CID_name:

			CID_name[ws["B" + str(i)].value] = [i]

	acme_wb = load_workbook("ACMEsetWithCIDs_new2.xlsx")
	ws_names = acme_wb["PubChemCID"]

	for i in range(2,483):

		CID = ws_names["C" + str(i)].value
		if CID in CID_name:

			CID_name[CID].append(str(ws_names["B" + str(i)].value).upper())

	PRR_wb = load_workbook("contingency_tables1.xlsx")
	PRR_ws = PRR_wb["Sheet2"]

	values = [v for k,v in CID_name.items()]

	names = [v[1] for v in values]
	indeces = [v[0] for v in values]

	for i in range(2, 1682, 21):

		acme_mol = str(PRR_ws["A" + str(i)].value).upper()

		if acme_mol not in names:

			continue

		index = indeces[names.index(acme_mol)]

		letters = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W"]

		for j in range(21):

			cell = letters[j] + str(index)
			print(cell)
			PRR = PRR_ws["G" + str(i+j)].value

			ws[cell].value = PRR

	wb.save("dimensionMatrix.xlsx")


# converts all of the PRRs into a binary 1 or 0; 1 for a positive signal (given by a PRR>=2), 0 for a negative
def binarize():

	wb = load_workbook("dimensionMatrix.xlsx")
	ws = wb["Sheet2"]

	for i in range(2, 82):

		letters = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W"]

		for letter in letters:

			cell = letter + str(i)

			PRR = ws[cell].value

			if PRR>=2:

				ws[cell].value = 1

			else:

				ws[cell].value = 0

	wb.save("dimensionMatrix.xlsx")


# fixes the training dataset to be sent to the log. regression/SVM
def createVector():

	# initialization 
	wb = load_workbook("dimensionMatrix.xlsx")
	ws1 = wb["S1"]
	ws2 = wb["S2"]
	ws5 = wb["S5"]
	ws6 = wb["S6"]

	unique_ccl = []
	for i in range(2,44758):

		ccl = int(ws1["C" + str(i)].value)

		if ccl not in unique_ccl:

			unique_ccl.append(ccl)

	unique_ccl = sorted(unique_ccl)

	cpd_signal = {} 

	letters = {"C": 558, "D": 348, "E": 47, "F": 103, "G": 63, "H": 14, "I": 13, "J": 7, "K": 167, "L": 58, "M": 64, "N": 58, "O": 273, "P": 10, "Q": 27, "R": 213, "S": 13, "T": 30, "U": 45, "V": 183, "W": 239}

	for i in range(2,82):

		cpd = int(ws2["A" + str(i)].value)

		if cpd not in cpd_signal:

			cpd_signal[cpd] = {}

		for letter in letters:

			cpd_signal[cpd][letter] = ws2[letter + str(i)].value

	cpd_zscores = {}

	for i in range(2,44758):

		cpd = int(ws1["A" + str(i)].value)

		if cpd not in cpd_zscores:

			cpd_zscores[cpd] = csr_matrix((1, 642)).toarray()[0]

		ccl = int(ws1["C" + str(i)].value)
		current = cpd_zscores[cpd]
		z_score = ws1["E" + str(i)].value

		current[unique_ccl.index(ccl)] = z_score

	return wb, unique_ccl, cpd_signal, cpd_zscores

# an attempt to use recursive feature elimination, however eventually scrapped for other methods
# example used is L1 regularization
def RFECV_select():
        
	wb, unique_ccl, cpd_signal, cpd_zscores = createVector()
	ws2 = wb["Sheet2"]


	letters = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W"]

	X = []
	Y = []


	for letter in letters:

		SOC = ws2[letter + "1"].value 

		print(SOC + "!!!!")

		for cpd in cpd_signal:

			Y.append(cpd_signal[cpd][letter])
			X.append(cpd_zscores[cpd])

		X = np.asarray(X)
		Y = np.asarray(Y)

		clf = linear_model.LogisticRegression()

		rfecv = RFECV(estimator = clf, step = 1, cv=cross_validation.LeaveOneOut(80), scoring = 'accuracy')

		rfecv.fit(X,Y)


		print("Optimal number of features for : %s %d" % (SOC, rfecv.n_features_))


		# # Plot number of features VS. cross-validation scores
		# plt.figure()
		# plt.xlabel("Number of features selected")
		# plt.ylabel("Cross validation score (nb of correct classifications)")
		# plt.plot(range(1, len(rfecv.grid_scores_) + 1), rfecv.grid_scores_)
		# plt.show()

		X = []
		Y = []
	

# contains my many attempts to modify the model
# the uncommented version is able to make the ROC curves for each SOC model based on the LOOCV results
# performed feature selection during each CV but took the most recurring features for the overall feature selection of the model
def logRegressionFirst():

	wb, unique_ccl, cpd_signal, cpd_zscores = createVector()

	letters = {"C": 558, "D": 348, "E": 47, "F": 103, "G": 63, "H": 14, "I": 13, "J": 7, "K": 167, "L": 58, "M": 64, "N": 58, "O": 273, "P": 10, "Q": 27, "R": 213, "S": 13, "T": 30, "U": 45, "V": 183, "W": 239}

	ws6 = wb["S6"]
	ws5 = wb["S5"]
	ws2 = wb["S2"]
	ws3 = wb["S3"]
	X = []
	Y = []
	ws_lines = wb["cell_lines"]
	
	ind = 2
	ind_new = 2

	# ~~~~~~~~~~~~~
	# logistic regression over each SOC
	for letter in letters:

		means = np.zeros(642)
		for cpd in cpd_signal:

			Y.append(cpd_signal[cpd][letter])
			X.append(cpd_zscores[cpd])

		X = np.asarray(X)
		Y = np.asarray(Y)

		clf = linear_model.LogisticRegression(C=1e5)

		# clf.fit(X, Y)
		cv = cross_validation.LeaveOneOut(80)

		# total = 0

		probas_ = []
		p_value_to_score = {}
		check = []
		for i, (train, test) in enumerate(cv):

			current = X[train]

			sel = feature_selection.GenericUnivariateSelect(mode='k_best', param = letters[letter])
			sel.fit(current, Y[train])

			current = sel.transform(current)
			means = np.add(means, sel.scores_)

			proba_ = clf.fit(current, Y[train]).predict_proba(sel.transform(X[test]))[0][1]
			probas_.append(proba_)
		

		means = np.multiply(means, 1/(len(cv)))

		cell_lines = {}

		for i in range(len(means)):

			cell_lines[unique_ccl[i]] = means[i]

		sorted_lines = [(k,v) for v,k in sorted([(v,k) for k,v in cell_lines.items()])]
		
		wanted = [sorted_lines[i] for i in range(len(sorted_lines)-1, len(sorted_lines)-1-letters[letter], -1)]
		
		# print([wanted[x][0] for x in range(10)])
		for ccl_i in wanted:

			ws6["A" + str(ind_new)].value = ws2[letter + "1"].value
			ws6["B" + str(ind_new)].value = ccl_i[0]
			ind_new += 1


		fpr, tpr, threshhold = roc_curve(Y, probas_)
		roc_auc = auc(fpr,tpr)

		print(str(ws2[letter + "1"].value) + ": " + str(roc_auc))
		plt.plot(fpr, tpr, lw=1, label='AUC = %0.2f' % (roc_auc))

		ws5["G" + str(ind)].value = roc_auc
		ws5["D" + str(ind)].value = letters[letter]
		ws5["A" + str(ind)].value = ws2[letter + "1"].value
		ind += 1

		plt.plot([0, 1], [0, 1], '--', color=(0.6, 0.6, 0.6), label='Luck')

		plt.xlabel('False Positive Rate')
		plt.ylabel('True Positive Rate')
		plt.title('ROC Curve for ' + str(ws2[letter + "1"].value))
		plt.legend(loc="lower right")
		plt.xlim([-0.05, 1.05])
		plt.ylim([-0.05, 1.05])
		plt.show()

		X = []
		Y = []

	# ~~~~~~~~~~~~~~~~~~~~~~~~~~

	# for letter in letters:

	# 	print(letter + "!!!")

	# 	times = np.zeros(642)
	# 	for cpd in cpd_signal:

	# 		Y.append(cpd_signal[cpd][letter])
	# 		X.append(cpd_zscores[cpd])

	# 	X = np.asarray(X)
	# 	Y = np.asarray(Y)

	# 	clf = linear_model.LogisticRegression(penalty = "l1", C=10)

	# 	# clf.fit(X, Y)
	# 	cv = cross_validation.LeaveOneOut(80)
	# 	probas_ = []

	# 	lengths = []


	# 	for i, (train, test) in enumerate(cv):

	# 		current_x = X[train][:]
	# 		current_y = Y[train][:]
	# 		# print(current)
	# 		# attempted to use feature selection, but wasn't sure if doing it right

	# 		model = feature_selection.SelectFromModel(clf)
	# 		model.fit(current_x, current_y)

	# 		chosen = model.get_support()
	# 		indeces_model = np.transpose(np.nonzero(chosen))
	# 		for j in range(indeces_model.shape[0]):

	# 			times[indeces_model[j][0]] += 1


	# 		new_current_x = model.transform(current_x)
	# 		length = new_current_x.shape[1]
	# 		lengths.append(length)

	# 		# print(new_current.shape)


	# 		# sel = feature_selection.GenericUnivariateSelect(feature_selection.f_regression, mode='k_best', param = letters[letter])
	# 		# sel.fit(current, Y[train])
	# 		# current = sel.transform(current)
	# 		# means = np.add(means, sel.pvalues_)
	# 		# print(sel.pvalues_[0])
	# 		# print(sel.scores_[0])

	# 		correct = linear_model.LogisticRegression(penalty = 'l1', C = 10).fit(new_current_x, current_y)

	# 		proba_ = correct.predict_proba(model.transform(X[test]))[0][1]
	# 		# # print(proba_)
	# 		probas_.append(proba_)

	# 		# # total += roc_auc

	# 	fpr, tpr, threshhold = roc_curve(Y, probas_)
	# 	roc_auc = auc(fpr,tpr)
	# 	n_features = int(math.floor(np.mean(lengths)))
	# 	SOC = ws2[letter + "1"].value
	# 	print(str(SOC) + ": " + str(roc_auc) + " " + str(n_features))

	# 	wanted = times.argsort()[-n_features:][::-1]
	# 	real_times = times[wanted]


	# 	ws3["A" + str(ind)].value = SOC
	# 	ws3["B" + str(ind)].value = roc_auc
	# 	ws3["F" + str(ind)].value = n_features
	# 	ind += 1

	# 	for i in range(n_features):

	# 		ws_lines["A" + str(ind_new)].value = SOC
	# 		ws_lines["B" + str(ind_new)].value = unique_ccl[wanted[i]]
	# 		ws_lines["C" + str(ind_new)].value = real_times[i]
	# 		ind_new += 1


	# 	plt.plot(fpr, tpr, lw=1, label='LOOCV Area = %0.2f' % (roc_auc))

	# 	# ws5["G" + str(ind)].value = roc_auc
	# 	# ind += 1

	# 	plt.plot([0, 1], [0, 1], '--', color=(0.6, 0.6, 0.6), label='Luck')

	# 	plt.xlabel('False Positive Rate')
	# 	plt.ylabel('True Positive Rate')
	# 	plt.title('ROC-LOOCV Curve for ' + str(SOC))
	# 	plt.legend(loc="lower right")
	# 	plt.xlim([-0.05, 1.05])
	# 	plt.ylim([-0.05, 1.05])
	# 	plt.show()
	# 	X = []
	# 	Y = []
	# 	print("~~~~~~~~~~~~~~~~~~~~")



	# 	# ~~~~~~~~TESTING STARTS HERE~~~~~~~~~~~~~~

	# 	letter = "J"
	# 	for cpd in cpd_signal:

	# 		Y.append(cpd_signal[cpd][letter])
	# 		X.append(cpd_zscores[cpd])

	# 	X = np.asarray(X)
	# 	Y = np.asarray(Y)

	# 	clf = linear_model.LogisticRegression(C=1e5)

	# 	# clf.fit(X, Y)
	# 	cv = cross_validation.LeaveOneOut(80)

	# 	# total = 0
	# 	mean_tpr = 0.0
	# 	mean_fpr = np.linspace(0, 1, 100)

	# 	probas_ = []

	# 	for i, (train, test) in enumerate(cv):

	# 		current = X[train]
	# 		# print(current)
	# 		# attempted to use feature selection, but wasn't sure if doing it right
	# 		sel = feature_selection.GenericUnivariateSelect(feature_selection.f_regression, mode='k_best', param = )
	# 		sel.fit(current, Y[train])
	# 		current = sel.transform(current)
	# 		proba_ = clf.fit(current, Y[train]).predict_proba(sel.transform(X[test]))[0][1]
	# 		# print(proba_)
	# 		probas_.append(proba_)
	# 		# total += roc_auc
			
	# 	# avg = total/4

	# 	# # print(ws2[letter+"1"].value + ": " + str(avg))


	# 	fpr, tpr, threshhold = roc_curve(Y, probas_)
	# 	roc_auc = auc(fpr,tpr)

	# 	print(str(ws2[letter + "1"].value) + ": " + str(roc_auc))
	# 	# plt.plot(fpr, tpr, lw=1, label='LOOCV Area = %0.2f' % (roc_auc))


	# 	# ws3["A" + str(ind)].value = ws2[letter + "1"].value
	# 	# ws3["B" + str(ind)].value = roc_auc
	# 	# ind += 1

	# 	# plt.plot([0, 1], [0, 1], '--', color=(0.6, 0.6, 0.6), label='Luck')
	# 	# plt.plot(mean_fpr, mean_tpr, 'k--',
	# #        label='Mean ROC (area = %0.2f)' % mean_auc, lw=2)

	# 	# plt.xlabel('False Positive Rate')
	# 	# plt.ylabel('True Positive Rate')
	# 	# plt.title('ROC-LOOCV Curve for ' + str(ws2[letter + "1"].value))
	# 	# plt.legend(loc="lower right")
	# 	# plt.xlim([-0.05, 1.05])
	# 	# plt.ylim([-0.05, 1.05])
	# 	# plt.show()

	# 	X = []
	# 	Y = []

	# 	#~~~~~~~~~~~TESTING ENDS HERE~~~~~~~~~~~~~~


	# support vector machine
	# ind = 2
	# for letter in letters:

	# 	for cpd in cpd_signal:

	# 		Y.append(cpd_signal[cpd][letter])
	# 		X.append(cpd_zscores[cpd])

	# 	X = np.asarray(X)
	# 	Y = np.asarray(Y)

	# 	clf = svm.SVC(kernel = 'linear', probability = True)
	# 	clf.fit(X, Y)
	# 	cv = cross_validation.StratifiedKFold(Y, n_folds = 4)
	# 	total = 0

	# 	for i, (train, test) in enumerate(cv):

	# 		# rfe = feature_selection.RFE(clf, 20)
	# 		probas_ = clf.fit(X[train], Y[train]).predict_proba(X[test])
	# 		fpr, tpr, threshhold = roc_curve(Y[test], probas_[:,1])
	# 		roc_auc = auc(fpr,tpr)
	# 		total += roc_auc

	# 	avg = total/len(cv)
	# 	print(ws2[letter+"1"].value + ": " + str(avg))

	# 	ws4["A" + str(ind)].value = ws2[letter + "1"].value
	# 	ws4["B" + str(ind)].value = avg

	# 	ind += 1

	# 	X = []
	# 	Y = []


	# wb.save("dimensionMatrix.xlsx")

# looked at how to choose which cell lines to take data from; since not every cell line had data for the drugs
def signalDistribution():

	wb = load_workbook("dimensionMatrix.xlsx")
	ws = wb["Sheet2"]

	letters = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W"]

	x = []
	for letter in letters:

		count = 0

		for i in range(2,82):

			cell = letter + str(i)

			if ws[cell].value == 1:

				count += 1

		x.append(count)

	n, bins, patches = plt.hist(x, 50, normed=1, facecolor='green', alpha=0.75)
	
	y = mlab.normpdf(bins, np.mean(x), np.std(x))

	print("Mean: " + str(np.mean(x)))
	print("Standard Deviation: " + str(np.std(x)))
	l = plt.plot(bins, y, 'r--', linewidth=1)

	plt.xlabel('Positive Drug Signals')
	plt.ylabel('Probability')
	plt.title('Histogram of Signal Distribution')
	plt.show()



# just playing around with the ACME dataset, trying to aggregate certain pieces of info
def cellLineIdentification():

	acme_wb = load_workbook("ACMEsetWithCIDs_new2.xlsx")
	ccl_ws = acme_wb["S2"]

	wb = load_workbook("dimensionMatrix.xlsx")
	ws6 = wb["cell_lines"]

	ccl_data = {}

	for i in range(2, 666):

		ccl_id = int(ccl_ws["A" + str(i)].value)
		ccl_name = ccl_ws["B" + str(i)].value
		culture_media = ccl_ws["C" + str(i)].value
		primary_site = ccl_ws["D" + str(i)].value
		histology = ccl_ws["E" + str(i)].value

		ccl_data[ccl_id] = [ccl_name, culture_media, primary_site, histology]

	for i in range(2, 732):

		ccl_id = int(ws6["B" + str(i)].value)
		data = ccl_data[ccl_id]

		ws6["D" + str(i)].value = data[0]
		ws6["E" + str(i)].value = data[1]
		ws6["F" + str(i)].value = data[2]
		ws6["G" + str(i)].value = data[3]

	wb.save("dimensionMatrix.xlsx")

# calculates standard error of AUC of ROC curves
def calculateError():

	wb = load_workbook('dimensionMatrix.xlsx')
	ws5 = wb["S3"]

	SOC_AUC = {}

	for i in range(2, 23):

		SOC = ws5["A" + str(i)].value
		AUC = ws5["B" + str(i)].value
		n_pos = ws5["F" + str(i)].value
		SOC_AUC[SOC] = [AUC, n_pos, i]

	for SOC in SOC_AUC:

		A = SOC_AUC[SOC][0]
		n_pos = SOC_AUC[SOC][1]
		ind = SOC_AUC[SOC][2]
		n_neg = 80 - n_pos
		Q1 = A/(2-A)
		Q2 = (2*(A**2))/(1+A)

		standard_error = math.sqrt((A*(1-A) + (n_pos-1)*(Q1-A**2) + (n_neg - 1)*(Q2 - A**2))/(n_pos*n_neg))
		print(SOC + ": " + str(standard_error))
		ws5["D" + str(ind)].value = standard_error


	wb.save('dimensionMatrix.xlsx')

# listed the mutations associated to the chosen cell lines of feature selection
def mutationCollect():

	wb = load_workbook('dimensionMatrix.xlsx')
	ws6 = wb["diff"]

	acme_wb = load_workbook("ACMEsetWithCIDs_new2.xlsx")
	mut_ws = acme_wb["S6"]

	ccl_mutation = {}

	for i in range(2,3123):

		d = mut_ws['D' + str(i)].value
		mutation = mut_ws['C' + str(i)].value
		cell_lines = [int(x) for x in d.split(";")]

		for ccl in cell_lines:

			if ccl not in ccl_mutation:

				ccl_mutation[ccl] = []

			ccl_mutation[ccl].append(mutation)

	print("Finished init!")

	for i in range(2,370):

		ccl = int(ws6["B" + str(i)].value)

		if ccl not in ccl_mutation:

			continue

		mutations = ccl_mutation[ccl][:]
		write = ""
		for j in range(len(mutations)):

			write += str(mutations[j]) + ";"

		write = write[:-1]

		ws6["H" + str(i)].value = write

	wb.save('dimensionMatrix.xlsx')

# showed a ratio of the number of times a SOC had a certain mutation
def SOC_mutation():

	wb = load_workbook('dimensionMatrix.xlsx')
	ws6 = wb["Sheet6"]
	ws7 = wb["Sheet7"]

	SOC_mutations = {}

	for i in range(2, 2535):

		SOC = ws6["A" + str(i)].value

		if SOC not in SOC_mutations:

			SOC_mutations[SOC] = {}

		current = SOC_mutations[SOC]
		muts = ws6["G" + str(i)].value

		if muts is not None:

			mutations = muts.split(";")

			for mutation in mutations:

				if mutation not in current:

					current[mutation] = 1

				else:

					current[mutation] += 1

	print("Done init! Now calculating totals...")
	for SOC in SOC_mutations:

		current = SOC_mutations[SOC]

		total = 0

		for mutation in current:

			count = current[mutation]
			total += count 

		for mutation in current:

			count = current[mutation]
			ratio = float(count)/total
			current[mutation] = ratio

	ind = 2
	for SOC in SOC_mutations:

		print(SOC + "!!!")

		current = SOC_mutations[SOC]

		ordered = [(k,v) for (v,k) in sorted([(v,k) for (k,v) in current.items()], reverse = True)]

		for k,v in ordered:

			ws7["A" + str(ind)].value = SOC
			ws7["B" + str(ind)].value = k
			ws7["C" + str(ind)].value = v
			ind += 1

	wb.save('dimensionMatrix.xlsx')

# same as previous, but instead looks to see how many times a mutation was associated with a certain SOC
def mutation_SOC():	

	wb = load_workbook("dimensionMatrix.xlsx")
	ws6 = wb["Sheet6"]
	ws8 = wb["Sheet8"]

	mutation_SOCs = {}

	for i in range(2, 2535):

		SOC = ws6["A" + str(i)].value
		muts = ws6["G" + str(i)].value

		if muts is not None:

			mutations = muts.split(";")

			for mutation in mutations:

				if mutation not in mutation_SOCs:

					mutation_SOCs[mutation] = {}

				current = mutation_SOCs[mutation]

				if SOC not in current:

					current[SOC] = 1

				else:

					current[SOC] += 1

	for mutation in mutation_SOCs:

		current = mutation_SOCs[mutation]

		total = 0

		for SOC in current:

			count = current[SOC]
			total += count

		for SOC in current:

			count = current[SOC]
			ratio = float(count)/total
			current[SOC] = ratio

	ind = 2
	for mutation in mutation_SOCs:

		current = mutation_SOCs[mutation]

		ordered = [(k,v) for (v,k) in sorted([(v,k) for (k,v) in current.items()], reverse = True)]

		for SOC, ratio in ordered:

			ws8["A" + str(ind)].value = mutation
			ws8["B" + str(ind)].value = SOC
			ws8["C" + str(ind)].value = ratio

			ind += 1

	wb.save("dimensionMatrix.xlsx")


# attempt to make graph to be used for presentation, but found a different way to do so in excel
def plotComparison():

	wb = load_workbook('dimensionMatrix.xlsx')
	ws5 = wb["Sheet5"]

	SOC_AUCs = {}

	for i in range(2, 23):

		if i == 16 or i == 18:

			continue

		SOC = ws5["A" + str(i)].value
		ours = ws5["B" + str(i)].value
		pouillot = ws5["C" + str(i)].value
		error = ws5["G" + str(i)].value
		SOC_AUCs[SOC] = [ours, pouillot, error]

	Y1 = []
	Y2 = []
	names = []

	for SOC in SOC_AUCs:

		Y1.append(SOC_AUCs[SOC][0])
		Y2.append(SOC_AUCs[SOC][1])
		names.append(SOC)

	names = tuple(names)
	N = 19

	ind = np.arange(N)
	width = 0.3

	fig = plt.figure()
	ax = fig.add_subplot(111)

	ours = ax.bar(ind, Y1, width, color = 'r')
	theirs = ax.bar(ind+width, Y2, width, color = 'b')

	ax.set_ylabel('AUC')
	ax.set_xticks(ind+width)
	ax.set_xticklabels(names, rotation='vertical')

	ax.legend( (ours[0], theirs[0]), ('r', 'b') )


	plt.show()

# applied the same techniques, however features were chosen by L1 regularization rather than K-best selection
def findDifferences():

	wb = load_workbook("dimensionMatrix.xlsx")
	lines_ws = wb["cell_lines"]
	ws6 = wb["S6"]
	diff = wb["diff"]

	L1_cell_lines = {}

	for i in range(2,732):

		SOC = lines_ws["A" + str(i)].value

		if SOC not in L1_cell_lines:

			L1_cell_lines[SOC] = {}

		ccl = lines_ws["B" + str(i)].value
		n = lines_ws["C" + str(i)].value
		name = lines_ws["D" + str(i)].value
		culture = lines_ws["E" + str(i)].value
		ccle_primary = lines_ws["F" + str(i)].value
		histology = lines_ws["G" + str(i)].value

		data = [n, name, culture, ccle_primary, histology]
		L1_cell_lines[SOC][ccl] = data

	orig_cell_lines = {}

	ind = 2

	for i in range(2, 2535):

		SOC = ws6["A" + str(i)].value
		if SOC not in orig_cell_lines:

			orig_cell_lines[SOC] = {}

		ccl = ws6["B" + str(i)].value

		orig_cell_lines[SOC][ccl] = ""

	for SOC in L1_cell_lines:

		current = L1_cell_lines[SOC]

		for ccl in current:

			if ccl not in orig_cell_lines[SOC]:

				data = current[ccl]

				diff["A" + str(ind)].value = SOC
				diff["B" + str(ind)].value = ccl
				diff["C" + str(ind)].value = data[0]
				diff["D" + str(ind)].value = data[1]
				diff["E" + str(ind)].value = data[2]
				diff["F" + str(ind)].value = data[3]
				diff["G" + str(ind)].value = data[4]
				ind += 1

	wb.save("dimensionMatrix.xlsx")


main()
