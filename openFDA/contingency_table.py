# Goal: calculate the PRRs for each SOC - ACME drug relationship

from __future__ import division
import simplejson as json
from collections import OrderedDict
from openpyxl import load_workbook, Workbook
from urllib2 import Request, urlopen, URLError


def main():

	partA()
	# partB()
	# partC()
	# partD()
	# PRR()

# count of reports associated WITH the ACME molecule AND SOC
def partA():

	file = open("Drug_ADE_ACME_Update_MAIN.json")
	json_data = json.load(file)
	file.close()

	ADE_wb = load_workbook("ADE_Database.xlsx")
	ADE_ws = ADE_wb["Sheet1"]
	# ADE_ws 22046

	acme_wb = load_workbook("acmeSynonyms.xlsx")
	acme_ws = acme_wb["Sheet1"]
	# ACME_ws 22997

	table_wb = load_workbook("contingency_tables2.xlsx")
	table_ws = table_wb["Sheet1"]
	index = 2

	acme_db = {}

	for i in range(2,22997):

		mol = str(acme_ws["A" + str(i)].value)

		if mol not in acme_db:

			acme_db[mol] = {}
			acme_db[mol][mol.upper()] = ""

		syn = str(acme_ws["C" + str(i)].value).upper()

		if syn not in acme_db[mol]:

			acme_db[mol][syn] = ""

	SOCs = {}

	for i in range(2, 22046):

		SOC = ADE_ws["A" + str(i)].value
		PT = str(ADE_ws["B" + str(i)].value).upper()
		terms = []

		if ADE_ws["C" + str(i)].value:

			classifies = str(ADE_ws["C" + str(i)].value).upper()
			terms = classifies.split("$")[:]

		if SOC not in SOCs:

			SOCs[SOC] = {}

		SOCs[SOC][PT] = ""

		for term in terms:

			SOCs[SOC][term] = ""

	print("Done with initialization.")

	num_found = 0
	for acme in acme_db:

		print(acme)

		n_soc = 0

		for SOC in SOCs:

			total = 0

			num_syn = 0

			syn_found = []
			for syn in acme_db[acme]:

				count = 0
				current = json_data.get(syn, {})

				if current == {}:

					# print(syn + " was not found!")
					continue

				num_syn += 1
				syn_found.append(syn)
				for ADE in current:

					real = ADE.replace("^", "'")

					if real in SOCs[SOC]:

						count += json_data[syn][ADE]

				total += count

			if total != 0:

				n_soc += 1
				print(acme + ":" + SOC + " " + str(num_syn) + " synonyms were found!")
				print(syn_found)

			# print(acme + " had a total of " + str(total) + " reports with " + str(SOC) + "!!")
			table_ws["A" + str(index)].value = acme
			table_ws["B" + str(index)].value = SOC
			table_ws["C" + str(index)].value = total
			index += 1

		if n_soc != 0:

			num_found += 1

	print("There were " + str(num_found) + " ACME molecules found in FAERS.")
	print("Saving...")
	table_wb.save("contingency_tables2.xlsx")
	table_wb = load_workbook("contingency_tables2.xlsx")
	table_ws = table_wb["Sheet1"]

# count of reports associated WITHOUT the ACME molecule but WITH the SOC
def partB():

	SOC_counts = {}
	file = open("SOC_counts_MAIN.txt")

	for line in file:

		real_line = line[:-1]

		parts = real_line.split("-")

		SOC_counts[parts[0]] = int(parts[1])

	wb = load_workbook("contingency_tables1.xlsx")
	ws = wb["Sheet1"]

	for i in range(2, 9830):

		SOC = ws["B" + str(i)].value
		ws["D" + str(i)].value = SOC_counts[SOC] - int(ws["C" + str(i)].value)

	wb.save("contingency_tables1.xlsx")

# count of reports associated WITH the ACME molecule but NOT the SOC
def partC():

	print("PART C")

	file = open("Drug_ADE_ACME_Update_MAIN.json")
	json_data = json.load(file)
	file.close()

	wb = load_workbook("contingency_tables1.xlsx")
	ws = wb["Sheet1"]
	# 9830

	acme_wb = load_workbook("acmeSynonyms.xlsx")
	acme_ws = acme_wb["Sheet1"]
	# 22997

	ADE_wb = load_workbook("ADE_Database.xlsx")
	ADE_ws = ADE_wb["Sheet1"]


	acme_syns = {}
	for i in range(2, 22997):

		acme = str(acme_ws["A" + str(i)].value)

		if acme not in acme_syns:

			acme_syns[acme] = [acme.upper()]

		syn = str(acme_ws["C" + str(i)].value).upper()

		if syn not in acme_syns[acme]:

			acme_syns[acme].append(syn)

	acme_counts = {}

	for acme in acme_syns:

		synonyms = acme_syns[acme]

		total = 0

		for syn in synonyms:

			count = 0

			current = json_data.get(syn, {})

			if current == {}:

				continue

			for ADE in current:

				count += current[ADE]

			total += count

		acme_counts[acme] = total

	for i in range(2, 9830):

		acme = str(ws["A" + str(i)].value)

		ws["E" + str(i)].value = acme_counts[acme] - int(ws["C"+str(i)].value)

	wb.save("contingency_tables1.xlsx")

# count of reports not associated with ACME NOR SOC
def partD():

	print("PART D")
	total = 0
	file = open("SOC_counts_MAIN.txt")

	for line in file:

		real_line = line[:-1]

		parts = real_line.split("-")

		total += int(parts[1])

	wb = load_workbook("contingency_tables1.xlsx")
	ws = wb["Sheet1"]

	print(total)

	for i in range(2, 9830):

		ws["F" + str(i)].value = total - int(ws["C" + str(i)].value) - int(ws["D" + str(i)].value) - int(ws["E" + str(i)].value)

	wb.save("contingency_tables1.xlsx")


def PRR():

	print("PRR")
	wb = load_workbook("contingency_tables1.xlsx")
	ws = wb["Sheet1"]
	for i in range(2, 9830):

		A = int(ws["C" + str(i)].value)
		B = int(ws["D" + str(i)].value)
		C = int(ws["E" + str(i)].value)
		D = int(ws["F" + str(i)].value)

		if A + C == 0:

			ws["G" + str(i)].value = 0

		else:


			print(str(A) + " " + str(B) + " " + str(C) + " " + str(D))
			one = (A/(A+C))
			two = (B/(B+D))
			ws["G" + str(i)].value = one/two


	wb.save("contingency_tables1.xlsx") 


main()
